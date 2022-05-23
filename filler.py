from constants import _DATA_BASE
import sqlite3
import openpyxl

def import_medicaments():
    db = sqlite3.connect(_DATA_BASE)
    cr = db.cursor()
    instruction = "create table if not exists Medicaments(ID integer, libelle text, forme integer)"
    cr.execute(instruction)
    wb = openpyxl.load_workbook("LISTE MEDICAMENTS.xlsx")
    sheet=wb['Sheet1']
    for i in range(2, 140, 1):
        cr.execute(f"insert into Medicaments(ID, libelle, forme) values ({int(str(sheet.cell(row=i, column=1).value) + str(sheet.cell(row = i, column = 2).value))}, \"{sheet.cell(row = i, column = 3).value}\", {sheet.cell(row = i, column = 4).value})")
    db.commit()
    db.close()

def import_formes():
    db = sqlite3.connect(_DATA_BASE)
    cr = db.cursor()
    instruction = "create table if not exists Formes(ID integer, libelle text, sigle text, dosage text, unite text)"
    cr.execute(instruction)
    wb = openpyxl.load_workbook("FORMES.xlsx")
    sheet=wb['Sheet1']
    for i in range(2, 2058, 1):
        instr = f"insert into Formes(ID, libelle, sigle, dosage, unite) values ({sheet.cell(row = i, column = 1).value}, \"{sheet.cell(row = i, column = 2).value}\", \"{sheet.cell(row = i, column = 3).value}\", \"{sheet.cell(row = i, column = 4).value}\", \"{sheet.cell(row = i, column = 5).value}\")"
        cr.execute(instr)

    db.commit()
    db.close()

def import_pathologies():
    db = sqlite3.connect(_DATA_BASE)
    cr = db.cursor()
    instruction = "create table if not exists Pathologies(ID integer,classe text, code text, maladie text)"
    cr.execute(instruction)
    wb = openpyxl.load_workbook("PATHOLOGIES.xlsx")
    sheet=wb['Sheet1']
    for i in range(2, 848, 1):
        instr = f"insert into Pathologies(ID, classe, code, maladie) values ({sheet.cell(row = i, column = 1).value}, \"{sheet.cell(row = i, column = 2).value[0]}\", \"{sheet.cell(row = i, column = 2).value[1:]}\", \"{sheet.cell(row = i, column = 3).value}\")"
        cr.execute(instr)
    db.commit()
    db.close()


def import_wilayas():
    db = sqlite3.connect(_DATA_BASE)
    cr = db.cursor()
    instruction = "create table if not exists Wilayas(Code integer, nom text)"
    cr.execute(instruction)
    wb = openpyxl.load_workbook("LISTE WILAYAS.xlsx")
    sheet=wb['Sheet1']
    for i in range(1, 59, 1):
        cr.execute(f"insert into Wilayas(Code, nom) values ({sheet.cell(row = i, column = 1).value}, \"{sheet.cell(row = i, column = 2).value}\")")
    db.commit()
    db.close()
