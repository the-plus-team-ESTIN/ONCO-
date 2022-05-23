import fpdf
from Patient_class import Patient
from Session_class import Seance
from Test_class import Test
import core
import openpyxl as xl
from datetime import date 
from math import ceil
import datetime as dt
from matplotlib.backends.backend_pdf import PdfPages
from pandas import DataFrame
import matplotlib.pyplot as plt

def printWeightPlot(main, file, PATIENT:Patient):
    program = main.core
    SEANCES = program._ALL_PATIENT_SESSIONS(PATIENT.ID)
    SEANCES.sort(key=lambda x: x.date_seance)
    WeightInformation = program.getPatientsWeightInformation(SEANCES)
    df1 = DataFrame(WeightInformation,columns=['Dates','Poids'])
    with PdfPages(file) as export_pdf:
        plt.plot(df1['Dates'], df1['Poids'], color='red', marker='o')
        plt.title('Changement du poids', fontsize=10)
        plt.xlabel('Dates', fontsize=8)
        plt.ylabel('Poids', fontsize=8)
        plt.grid(True)
        export_pdf.savefig()
        plt.close()


def printStatistics(file, program, date1, date2, unit):

    titles = ["N°", "Pathologies", "Homme AC", "Homme NC", "Femme AC", "Femme NC", "Total AC", "Total NC", "Total GN"]
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Statistiques"
    i = 1
    for title in titles:
        sheet.cell(row = 15, column = i).value = title
        i += 1
    i = 16
    for path in program.getAllPathologies():
        result = program.stat_pathologie_partiel(date1, date2, unit, path[0])
        sheet.cell(row = i, column = 1).value = path[0]
        sheet.cell(row = i, column = 2).value = path[1] + " " + path[2] + " " + path[3]
        sheet.cell(row = i, column = 3).value = result[0]
        sheet.cell(row = i, column = 4).value = result[1]
        sheet.cell(row = i, column = 5).value = result[2]
        sheet.cell(row = i, column = 6).value = result[3]
        sheet.cell(row = i, column = 7).value = result[4]
        sheet.cell(row = i, column = 8).value = result[5]
        sheet.cell(row = i, column = 9).value = result[6]
        i += 1

    totalACHomme = program.stat_totalACHomme(date1, unit)
    totalNCHomme = program.stat_totalNCHomme(date1, date2, unit)
    totalACFemme = program.stat_totalACFemme(date1, unit)
    totalNCFemme = program.stat_totalNCFemme(date1, date2, unit)
    totalAC = totalACHomme + totalACFemme
    totalNC = totalNCHomme + totalNCFemme
    totalTG = totalNC + totalAC
    sheet.cell(row = 863, column = 1).value = "Total Homme AC"
    sheet.cell(row = 863, column = 2).value = totalACHomme
    sheet.cell(row = 864, column = 1).value = "Total Homme NC"
    sheet.cell(row = 864, column = 2).value = totalNCHomme
    sheet.cell(row = 865, column = 1).value = "Total Femme AC"
    sheet.cell(row = 865, column = 2).value = totalACFemme
    sheet.cell(row = 866, column = 1).value = "Total Femme NC"
    sheet.cell(row = 866, column = 2).value = totalNCFemme
    sheet.cell(row = 867, column = 1).value = "Total AC"
    sheet.cell(row = 867, column = 2).value = totalAC
    sheet.cell(row = 868, column = 1).value = "Total NC"
    sheet.cell(row = 868, column = 2).value = totalNC
    sheet.cell(row = 869, column = 1).value = "Total TG"
    sheet.cell(row = 869, column = 2).value = totalTG
    #/filename = "Statistiques" + str(dt.date.today()) + ".xlsx"
    wb.save(filename= file)


def printListPatients(file, PAT_LIST):
    titles = ["N° du dossier", "Nom", "Prénom", "Pathologies", "Unité", "date d'entrée"]
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Liste des patients"

    i = 1
    for title in titles:
        sheet.cell(row = 10, column = i).value = title
        i += 1

    i = 11
    for patient in PAT_LIST:
    # for j in range(5):
        sheet.cell(row = i, column = 1).value = patient.ID
        sheet.cell(row = i, column = 2).value = patient.nom
        sheet.cell(row = i, column = 3).value = patient.prenom
        sheet.cell(row = i, column = 4).value = patient.pathologies
        sheet.cell(row = i, column = 5).value = patient.UNITE()
        sheet.cell(row = i, column = 6).value = patient.date_entree
        print(patient)
        i += 1
    #filename="ListeDesPatients" + + str(dt.date.today()) + "".xslx"
    
    wb.save(file)

def printListDoctors(file, DOC_LIST):

    titles = ["Nom", "Prénom", "Unité", "Grade"]
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Liste des medecins"

    i = 1
    for title in titles:
        sheet.cell(row = 10, column = i).value = title
        i += 1

    i = 11
    for med in DOC_LIST:
    # for j in range(5):
        sheet.cell(row = i, column = 1).value = med.nom
        sheet.cell(row = i, column = 2).value = med.prenom
        sheet.cell(row = i, column = 3).value = med.UNITE()
        sheet.cell(row = i, column = 4).value = med.grade
        print(med)
        i += 1
    #filename="ListeDesMedecins" + str(dt.date.today()) + "".xslx"
    wb.save(filename = file)

def printTodaysSessions(main, file, LIST_TODAY):
    program = main.core
    titles = ["N° dossier", "Nom", "Prénom", "Unité", "Type de traitement"]
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Liste d'aujourd'hui"

    i = 1
    for title in titles:
        sheet.cell(row = 10, column = i).value = title
        i += 1
    print(LIST_TODAY)
    i = 11
    for seance in LIST_TODAY:
        patient = program._SELECT_PATIENT(seance.ID_patient)
        sheet.cell(row = i, column = 1).value = str(patient[0].ID)[:-2]+"/"+str(patient[0].ID)[-2:]
        sheet.cell(row = i, column = 2).value = patient[0].nom
        sheet.cell(row = i, column = 3).value = patient[0].prenom
        sheet.cell(row = i, column = 4).value = patient[0].UNITE()
        sheet.cell(row = i, column = 5).value = seance.TYPE_TRAITEMENT()
        i += 1
    #filename="ListeAujourdhui" + str(dt.date.today()) + "".xslx"
    wb.save(filename = file)


def NHeader(pdf, wd, hdr, nl, frm):
    pdf.set_font('Times', frm, 13)
    pdf.cell(wd, 7, hdr, 0, nl, 'C')

def Titre(pdf, ttr):
    pdf.set_font('Times', 'B', 20)
    pdf.cell(0, 7, ttr, 0, 1, 'C')

def STitre(pdf, wd, sttr, nl, al, frm):
    pdf.set_font('Times', frm, 13)
    pdf.cell(wd, 10, sttr, 0, nl, al)

def NTitre(pdf, wd, sttr, nl, al, frm):
    pdf.set_font('Times', frm, 13)
    pdf.cell(wd, 10, sttr, 1, nl, al)




def printPatient(program, file, PATIENT, LISTE_SEANCE, LISTE_TESTS):
    pdf = fpdf.FPDF()
    pdf.set_margins(top = 15, left=15, right = 15)
    pdf.add_page()
    NHeader(pdf, 0, "WILAYA DE BEJAIA", 1, 'B')
    NHeader(pdf, 0, "ETABLISSEMENT PUBLIC HOSPITALIER D'AMIZOUR", 1, 'B')
    NHeader(pdf, 0, "HOPITAL BEN-MERAD EL-MEKI", 1, 'BU')
    pdf.ln(7)
    Titre(pdf, "DOSSIER MEDICAL")
    pdf.ln(10)
    STitre(pdf, 15, 'Nom :', 0, 'L','BU')
    STitre(pdf, 50, PATIENT.nom, 0, 'C','')
    STitre(pdf, 81, 'Numéro du dossier :', 0, 'R', 'BU')
    STitre(pdf, 35, str(PATIENT.ID), 1, 'C', '')
    STitre(pdf, 22, 'Prénom :', 0, 'L','BU')
    STitre(pdf, 39, PATIENT.prenom, 0, 'C','')
    STitre(pdf, 74, 'Date d\'entrée :', 0, 'R', 'BU')
    STitre(pdf, 35, PATIENT.date_entree.isoformat(), 1, 'C', '')
    pdf.ln(5)
    STitre(pdf, 74, 'Pathologies :', 1, 'L', 'BU')
    paths = program.getPathologies(PATIENT)
    print(paths)
    if paths != [""]:
        for path in paths:
            double = ceil(len(path) / 87)
            for i in range(double):
                STitre(pdf, 0, path[(i * 88):((i + 1) * 88)], 1, 'L','')
        pdf.ln(5)
    STitre(pdf, 40, 'Date de naissance :', 0, 'L','BU')
    STitre(pdf, 45, PATIENT.date_naissance.isoformat(), 1, 'C','')
    STitre(pdf, 16, 'Sexe :', 0, 'L','BU')
    STitre(pdf, 30, PATIENT.SEXE(), 1, 'C','')
    STitre(pdf, 16, 'Age :', 0, 'L','BU')
    STitre(pdf, 30, str(PATIENT.AGE()), 1, 'C','')
    STitre(pdf, 16, 'Taille :', 0, 'L','BU')
    STitre(pdf, 30, str(PATIENT.taille) + " cm", 1, 'C','')
    STitre(pdf, 24, 'Domicile :', 0, 'L','BU')
    wilaya = program.getWilaya(PATIENT)
    STitre(pdf, 45, (PATIENT.commune + ", " + wilaya), 1, 'C','')
    STitre(pdf, 32, 'N° téléphone :', 0, 'L','BU')
    STitre(pdf, 40, str(PATIENT.tel_patient), 1, 'C','')
    STitre(pdf, 49, 'N° téléphone altérnatif :', 0, 'R', 'BU')
    STitre(pdf, 40, str(PATIENT.tel_alt), 1, 'C', '')
    # STitre(pdf, 0, 'Evolution du poids :', 1, 'R', 'BU')
    # # pdf.add_page()
    # fonction to print all the sessions in 
    pdf.add_page()
    Titre(pdf, "Liste des séances et des testes :")
    pdf.ln(5)
    nb_seance = 0
    nb_test = 0
    all_list:list = LISTE_SEANCE + LISTE_TESTS
    all_list.sort(key = lambda  x: x.date_seance if type(x) is Seance else x.date_test )
    for i in all_list:
        if (type(i)is Test):
            nb_test +=1
            STitre(pdf, 0, 'Test N° ' + str(nb_test) + " :", 1, 'C','BU')
            informationTest(pdf, i)
        else :
            nb_seance +=1
            STitre(pdf, 0, 'Séance N° ' + str(nb_seance) + " :", 1, 'C','BU')
            informationSeance(pdf, i)
    #filename=PatientNom + '_' + PatientPrenom + str(dt.date.today()) + ".pdf"
    pdf.output(file, 'F')

def informationSeance(pdf, SEANCE:Seance):
    program = core.Core()
    STitre(pdf, 50, 'Date de la séance :', 0, 'L','BU')
    STitre(pdf, 50, str(SEANCE.date_seance), 1, 'L','')
    STitre(pdf, 50, 'Type du traitement :', 0, 'L','BU')
    STitre(pdf, 50, SEANCE.TYPE_TRAITEMENT(), 1, 'L','')
    STitre(pdf, 50, 'Médicaments :', 1, 'L','BU')
    meds = program.getMedicaments(SEANCE)
    for med in meds:
        double = ceil(len(med) / 87)
        for i in range(double):
            STitre(pdf, 0, med[(i * 88):((i + 1) * 88)], 1, 'L','')
    STitre(pdf, 50, 'Poids du patient :', 0, 'L','BU')
    STitre(pdf, 50, str(SEANCE.poids) + " Kg", 1, 'L','')
    STitre(pdf, 50, 'Plus de détails :', 1, 'L','BU')
    double = ceil(len(SEANCE.details) / 87)
    for i in range(double):
        STitre(pdf, 0, SEANCE.details[(i * 88):((i + 1) * 88)], 1, 'L','')
    pdf.ln(5)

def informationTest(pdf, TEST:Test):
    STitre(pdf, 50, 'Type du test :', 0, 'L','BU')
    double = ceil(len(TEST.type) / 63)
    for i in range(double):
        STitre(pdf, 0, TEST.type[(i * 64):((i + 1) * 64)], 1, 'L','')
    STitre(pdf, 50, 'Date du Test :', 0, 'L','BU')
    STitre(pdf, 50, str(TEST.date_test), 1, 'L','')
    STitre(pdf, 50, 'Plus de détails :', 1, 'L','BU')
    double = ceil(len(TEST.details) / 87)
    for i in range(double):
        STitre(pdf, 0, TEST.details[(i * 88):((i + 1) * 88)], 1, 'L','')
    pdf.ln(5)

def printSeance(file, SEANCE, PATIENT):
    program = core.Core()
    pdf = fpdf.FPDF()
    pdf.set_margins(top = 15, left=15, right = 15)
    pdf.add_page()
    NHeader(pdf, 0, "WILAYA DE BEJAIA", 1, 'B')
    NHeader(pdf, 0, "ETABLISSEMENT PUBLIC HOSPITALIER D'AMIZOUR", 1, 'B')
    NHeader(pdf, 0, "HOPITAL BEN-MERAD EL-MEKI", 1, 'BU')
    pdf.ln(7)
    Titre(pdf, "Rapport de la séance")
    pdf.ln(10)
    STitre(pdf, 0, 'Numéro du dossier :', 1, 'C', 'BU')
    pdf.ln(3)
    STitre(pdf, 0, str(PATIENT.ID), 1, 'C', '')
    pdf.ln(7)
    STitre(pdf, 50, 'Date de la séance :', 0, 'L','BU')
    STitre(pdf, 50, str(SEANCE.date_seance), 1, 'L','')
    STitre(pdf, 50, 'Nom du patient:', 0, 'L','BU')
    STitre(pdf, 35, PATIENT.nom, 1, 'L', '')
    STitre(pdf, 50, 'Prénom du patient:', 0, 'L','BU')
    STitre(pdf, 50, PATIENT.prenom, 1, 'L','')
    STitre(pdf, 50, 'Unité :', 0, 'L','BU')
    STitre(pdf, 50, PATIENT.UNITE(), 1, 'L','')
    STitre(pdf, 50, 'Type du traitement :', 0, 'L','BU')
    STitre(pdf, 50, SEANCE.TYPE_TRAITEMENT(), 1, 'L','')
    STitre(pdf, 50, 'Médicaments :', 1, 'L','BU')
    meds = program.getMedicaments(SEANCE)
    for med in meds:
        double = ceil(len(med) / 87)
        for i in range(double):
            STitre(pdf, 0, med[(i * 88):((i + 1) * 88)], 1, 'L','')
    STitre(pdf, 50, 'Poids du patient :', 0, 'L','BU')
    STitre(pdf, 50, str(SEANCE.poids) + " Kg", 1, 'L','')
    pdf.ln(10)
    STitre(pdf, 50, 'Plus de détails :', 1, 'L','BU')
    STitre(pdf, 50, SEANCE.details, 1, 'L','')
    double = ceil(len(SEANCE.details) / 87)
    for i in range(double):
        STitre(pdf, 0, SEANCE.details[(i * 88):((i + 1) * 88)], 1, 'L','')
    #filename = PatientNom + "_" + PatientPrenom + "_seance" + str(dt.date.today()) + ".pdf"
    pdf.output((file + PATIENT.nom + "_" + PATIENT.prenom + "_Seance.pdf"), 'F')

